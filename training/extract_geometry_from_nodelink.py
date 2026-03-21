#!/usr/bin/env python3
"""
Extract geometry data from standard node-link SHP and generate training data.

Matches with speed CSV link IDs to extract:
- Actual speed limit (MAX_SPD)
- Number of lanes (LANES)
- Road rank (ROAD_RANK)
- Link length (LENGTH)
- Curve radius (calculated from geometry)

Usage:
    python -m training.extract_geometry_from_nodelink
"""

import json
import math
import os
import shapefile

SHP_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "raw_data", "[2026-01-13]NODELINKDATA", "MOCT_LINK"
)
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "data")

# ROAD_RANK code mapping
ROAD_RANK_MAP = {
    "101": "고속도로",
    "102": "도시고속도로",
    "103": "일반국도",
    "104": "특별광역시도",
    "105": "국가지원지방도",
    "106": "지방도",
    "107": "시군구도",
    "108": "기타",
}


def latlon_to_meters(lat1, lon1, lat2, lon2):
    """Calculate distance between two lat/lon points (Haversine formula, in meters)."""
    R = 6371000
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))


def calculate_curve_radius(points):
    """Calculate minimum curve radius from road shape coordinates (in meters).

    Computes curvature per segment using the circumscribed circle radius of 3 points,
    and returns the minimum value.
    Converts lat/lon coordinates to local meter coordinates for calculation.
    """
    if len(points) < 3:
        return None  # Straight line (2 points)

    # Convert lat/lon to local meter coordinates (relative to first point)
    ref_lat, ref_lon = points[0][1], points[0][0]
    m_per_deg_lat = 111320
    m_per_deg_lon = 111320 * math.cos(math.radians(ref_lat))

    local_pts = []
    for lon, lat in points:
        x = (lon - ref_lon) * m_per_deg_lon
        y = (lat - ref_lat) * m_per_deg_lat
        local_pts.append((x, y))

    radii = []
    for i in range(1, len(local_pts) - 1):
        x1, y1 = local_pts[i-1]
        x2, y2 = local_pts[i]
        x3, y3 = local_pts[i+1]

        # Circumscribed circle radius of three points
        a = math.dist((x1, y1), (x2, y2))
        b = math.dist((x2, y2), (x3, y3))
        c = math.dist((x1, y1), (x3, y3))

        # Skip segments that are too short
        if min(a, b, c) < 5:
            continue

        s = (a + b + c) / 2
        area_sq = s * (s-a) * (s-b) * (s-c)
        if area_sq <= 0:
            continue  # Collinear

        area = math.sqrt(area_sq)
        if area < 0.01:
            continue

        R = (a * b * c) / (4 * area)

        # Exclude unrealistically large values (nearly straight)
        if R < 50000:
            radii.append(R)

    if not radii:
        return None  # Nearly straight

    return int(min(radii))


def classify_curve(radius_m, speed_limit):
    """Convert curve radius to a human-readable description."""
    if radius_m is None:
        return "직선"

    min_r = {120: 710, 100: 460, 80: 280, 60: 150, 50: 90, 40: 60, 30: 30, 20: 15}
    min_for_speed = min_r.get(speed_limit, 100)

    ratio = radius_m / min_for_speed if min_for_speed > 0 else 10

    if ratio < 1.0:
        return "급커브"
    elif ratio < 1.5:
        return "커브"
    elif ratio < 3.0:
        return "완만한 곡선"
    elif ratio < 8.0:
        return "아주 완만한 곡선"
    else:
        return "거의 직선"


def extract_seoul_links():
    """Extract geometry of Seoul links from the SHP file."""
    print(f"Loading SHP file: {SHP_PATH}")
    sf = shapefile.Reader(SHP_PATH, encoding='euc-kr')
    print(f"  Total records: {len(sf)}")

    seoul_links = {}
    for srec in sf.iterShapeRecords():
        rec = srec.record
        lid = str(rec['LINK_ID'])

        # Seoul (code 11xx)
        if not lid.startswith('11'):
            continue

        shape = srec.shape
        pts = shape.points
        radius = calculate_curve_radius(pts)

        seoul_links[lid] = {
            "link_id": lid,
            "road_name": rec['ROAD_NAME'].strip() if rec['ROAD_NAME'] else "",
            "lanes": int(rec['LANES']) if rec['LANES'] else 1,
            "max_spd": int(rec['MAX_SPD']) if rec['MAX_SPD'] else 50,
            "road_rank": ROAD_RANK_MAP.get(str(rec['ROAD_RANK']), str(rec['ROAD_RANK'])),
            "road_rank_code": str(rec['ROAD_RANK']),
            "length_m": round(float(rec['LENGTH']), 1) if rec['LENGTH'] else 0,
            "num_points": len(pts),
            "min_curve_radius_m": radius,
            "curve_desc": classify_curve(radius, int(rec['MAX_SPD']) if rec['MAX_SPD'] else 50),
        }

    print(f"  Seoul links: {len(seoul_links)}")
    return seoul_links


def merge_with_speed_data(seoul_links: dict) -> dict:
    """Match with speed CSV link IDs and add geometry to existing training data."""
    import openpyxl
    import shutil

    speed_file = None
    raw_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "raw_data")
    import unicodedata
    for f in os.listdir(raw_dir):
        fn = unicodedata.normalize("NFC", f)
        if "속도" in fn or "통행속도" in fn:
            speed_file = os.path.join(raw_dir, f)
            break

    if not speed_file:
        print("  Speed CSV not found")
        return {}

    tmp = speed_file + ".tmp2.xlsx"
    shutil.copy2(speed_file, tmp)
    wb = openpyxl.load_workbook(tmp, read_only=True)
    ws = wb[wb.sheetnames[0]]

    matched = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row[3]:
            continue

        lid = str(int(row[3]))
        road_name = str(row[2]) if row[2] else ""

        if lid in seoul_links:
            geo = seoul_links[lid]
            key = f"{lid}_{road_name}"
            if key not in matched:
                matched[key] = {
                    "link_id": lid,
                    "road_name": road_name,
                    "lanes_nodelink": geo["lanes"],
                    "lanes_speed_csv": int(row[8]) if row[8] else None,
                    "max_spd": geo["max_spd"],
                    "road_rank": geo["road_rank"],
                    "length_m": geo["length_m"],
                    "min_curve_radius_m": geo["min_curve_radius_m"],
                    "curve_desc": geo["curve_desc"],
                    "road_type_speed_csv": str(row[9]) if row[9] else "",
                    "area_type": str(row[10]) if row[10] else "",
                }

    wb.close()
    os.remove(tmp)

    print(f"  Speed CSV matched: {len(matched)} (unique link x road name)")
    return matched


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("=" * 50)
    print("  Standard node-link geometry extraction")
    print("=" * 50)

    # 1) Extract Seoul links
    seoul_links = extract_seoul_links()

    # 2) Match with speed CSV
    merged = merge_with_speed_data(seoul_links)

    # 3) Statistics
    curves = [v for v in merged.values() if v["min_curve_radius_m"] is not None]
    straights = [v for v in merged.values() if v["min_curve_radius_m"] is None]
    print(f"\n  Curved links: {len(curves)}")
    print(f"  Straight links: {len(straights)}")

    if curves:
        radii = [v["min_curve_radius_m"] for v in curves]
        print(f"  Curve radius range: {min(radii)}m ~ {max(radii)}m")
        print(f"  Curve radius median: {sorted(radii)[len(radii)//2]}m")

    # Curve type distribution
    from collections import Counter
    curve_dist = Counter(v["curve_desc"] for v in merged.values())
    print(f"\n  Curve type distribution:")
    for desc, cnt in curve_dist.most_common():
        print(f"    {desc}: {cnt}")

    # Speed limit distribution
    spd_dist = Counter(v["max_spd"] for v in merged.values())
    print(f"\n  Speed limit distribution:")
    for spd, cnt in sorted(spd_dist.items()):
        print(f"    {spd}km/h: {cnt}")

    # 4) Save
    filepath = os.path.join(OUTPUT_DIR, "nodelink_geometry.jsonl")
    with open(filepath, "w", encoding="utf-8") as f:
        for v in merged.values():
            f.write(json.dumps(v, ensure_ascii=False) + "\n")
    print(f"\n  Saved: {filepath} ({len(merged)} records)")

    # Samples
    print(f"\n  Samples:")
    import random
    random.seed(42)
    for v in random.sample(list(merged.values()), min(5, len(merged))):
        print(f"    {v['road_name']}(ID:{v['link_id']}): "
              f"{v['lanes_nodelink']} lanes, {v['max_spd']}km/h, "
              f"{v['length_m']}m, R={v['min_curve_radius_m']}m ({v['curve_desc']})")


if __name__ == "__main__":
    main()
