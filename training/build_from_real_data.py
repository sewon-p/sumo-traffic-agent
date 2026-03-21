#!/usr/bin/env python3
"""
Generates fine-tuning JSONL from real data (Seoul speed/volume CSV).

Input:
  - Seoul vehicle travel speed (xlsx/csv): speed by road and time period + road class + lanes
  - Seoul traffic volume survey (xlsx/csv): volume by location and time period + coordinates

Output:
  - training/data/train_real_openai.jsonl: For OpenAI fine-tuning
  - training/data/train_real_gemini.jsonl: For Gemini tuning

Usage:
    python -m training.build_from_real_data
"""

import json
import os
import random
import openpyxl
from collections import defaultdict

RAW_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "raw_data")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "data")

SYSTEM_PROMPT = (
    "너는 교통공학 전문가이자 SUMO 시뮬레이션 엔지니어다. "
    "사용자가 도로/교통 상황을 설명하면, SUMO 시뮬레이션에 필요한 파라미터를 JSON으로만 반환한다. "
    "반드시 아래 8개 필드를 모두 숫자로 채워라. 절대 빈 값이나 문자열 '-'를 쓰지 마라.\n\n"
    "출력 형식:\n"
    '{"speed_kmh": 숫자, "volume_vph": 숫자, "lanes": 편도차로수, '
    '"speed_limit_kmh": 숫자, "sigma": 0~1사이, "tau": 0.5~3사이, '
    '"avg_block_m": 교차로간격(m), "reasoning": "판단근거"}'
)

TIME_LABELS = {
    (7, 9): "출근시간",
    (9, 12): "오전",
    (12, 14): "점심시간",
    (14, 17): "오후",
    (17, 20): "퇴근시간",
    (20, 23): "야간",
    (23, 7): "심야",
}


def get_time_label(hour: int) -> str:
    """Return the time-of-day label for the given hour."""
    for (h1, h2), label in TIME_LABELS.items():
        if h1 <= h2:
            if h1 <= hour < h2:
                return label
        else:  # Late night (23~7)
            if hour >= h1 or hour < h2:
                return label
    return "기타"


def load_speed_data(filepath: str) -> dict:
    """Aggregate speed data by road and time period (31-day average)."""
    print(f"Loading speed data: {filepath}")

    # Copy to xlsx for reading
    tmp = filepath + ".tmp.xlsx"
    import shutil
    shutil.copy2(filepath, tmp)

    wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Parse headers
    headers = None
    hour_cols = {}  # {hour: column_index}

    # Collect speeds by road and time period
    # key: (road_name, lanes, functional_type, urban/suburban) -> {hour: [speed_values]}
    road_speeds = defaultdict(lambda: defaultdict(list))
    road_info = {}  # (road_name) -> {lanes, functional_type, ...}

    row_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cols = list(row)
        if i == 0:
            headers = cols
            # Find time period columns: ~01시, ~02시, ... ~24시
            for j, h in enumerate(headers):
                if h and str(h).startswith("~") and "시" in str(h):
                    hour = int(str(h).replace("~", "").replace("시", ""))
                    hour_cols[hour] = j
            print(f"  Time period columns: {sorted(hour_cols.keys())}")
            continue

        if not cols[0]:
            continue

        road_name = str(cols[2])  # Road name
        lanes = cols[8]  # Number of lanes
        road_type = str(cols[9]) if cols[9] else ""  # Functional type
        area_type = str(cols[10]) if cols[10] else ""  # Urban/suburban
        day_of_week = str(cols[1])  # Day of week

        link_id = str(int(cols[3])) if cols[3] else ""  # Link ID
        road_key = road_name

        if road_key not in road_info:
            road_info[road_key] = {
                "lanes": lanes,
                "lanes_list": [lanes] if lanes else [],
                "road_type": road_type,
                "area_type": area_type,
                "link_ids": set(),
            }
        else:
            if lanes:
                road_info[road_key]["lanes_list"].append(lanes)
        road_info[road_key]["link_ids"].add(link_id)

        for hour, col_idx in hour_cols.items():
            speed = cols[col_idx]
            if speed is not None and isinstance(speed, (int, float)) and speed > 0:
                road_speeds[road_key][hour].append(float(speed))

        row_count += 1

    wb.close()
    os.remove(tmp)

    # Calculate averages and correct lane count using mode
    result = {}
    for road, hours in road_speeds.items():
        info = road_info.get(road, {})
        lanes_list = info.get("lanes_list", [])
        if lanes_list:
            from collections import Counter
            info["lanes"] = Counter(lanes_list).most_common(1)[0][0]
        result[road] = {
            "info": info,
            "hourly_speed": {},
        }
        for hour, speeds in hours.items():
            result[road]["hourly_speed"][hour] = round(sum(speeds) / len(speeds), 1)

    print(f"  Roads: {len(result)}, total rows: {row_count}")
    return result


def load_volume_data(filepath: str) -> dict:
    """Aggregate volume data by location and time period."""
    print(f"Loading volume data: {filepath}")

    tmp = filepath + ".tmp.xlsx"
    import shutil
    shutil.copy2(filepath, tmp)

    wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)

    # Data sheet
    ws = wb['2025년 10월']

    # Coordinates sheet
    ws_coord = wb['수집지점 주소 및 좌표']
    coords = {}
    for i, row in enumerate(ws_coord.iter_rows(values_only=True)):
        if i == 0:
            continue
        cols = list(row)
        if cols[0]:
            coords[str(cols[0])] = {
                "name": str(cols[1]) if cols[1] else "",
                "lat": float(cols[3]) if cols[3] else None,
                "lon": float(cols[4]) if cols[4] else None,
                "address": str(cols[5]) if cols[5] else "",
            }

    # Volume aggregation
    # key: (point_name) -> {hour: [volume_values]}
    point_volumes = defaultdict(lambda: defaultdict(list))
    point_info = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cols = list(row)
        if i == 0:
            continue

        if not cols[0]:
            continue

        point_name = str(cols[3])  # Point name
        point_id = str(cols[4])  # Point ID
        direction = str(cols[5])  # Direction
        day_type = str(cols[2])  # Weekday/weekend

        point_key = point_name

        if point_key not in point_info:
            point_info[point_key] = {
                "point_id": point_id,
                "direction": direction,
                "coord": coords.get(point_id, {}),
            }

        # Hourly volume 0~23 (columns 7~30)
        for hour in range(24):
            col_idx = 7 + hour
            if col_idx < len(cols):
                vol = cols[col_idx]
                if vol is not None and isinstance(vol, (int, float)) and vol >= 0:
                    point_volumes[point_key][hour].append(int(vol))

    wb.close()
    os.remove(tmp)

    # Calculate averages
    result = {}
    for point, hours in point_volumes.items():
        result[point] = {
            "info": point_info.get(point, {}),
            "hourly_volume": {},
        }
        for hour, volumes in hours.items():
            result[point]["hourly_volume"][hour] = int(sum(volumes) / len(volumes))

    print(f"  Points: {len(result)}")
    return result


def estimate_cf_params(speed_kmh: float, speed_limit: int, road_type: str) -> dict:
    """Reverse-estimate car-following parameters from speed/volume.

    Based on the Greenshields model: the lower the speed relative to free-flow
    speed, the higher the V/C ratio.
    speed_ratio = speed / free_speed
    V/C ~ 1 - speed_ratio (linear approximation)
    """
    if speed_limit == 0:
        speed_limit = 50

    free_speed = speed_limit * 0.9
    speed_ratio = speed_kmh / free_speed if free_speed > 0 else 0.5
    speed_ratio = min(max(speed_ratio, 0.05), 1.2)  # Clipping

    # Greenshields approximation: V/C = 1 - speed_ratio
    # Correction: even at full free-flow (ratio=1.0), V/C is about 0.1
    vc = max(0.05, min(1.0 - speed_ratio * 0.85, 0.98))
    # speed=15, limit=50 -> ratio=0.33 -> vc=0.72
    # speed=26, limit=50 -> ratio=0.58 -> vc=0.51
    # speed=40, limit=50 -> ratio=0.89 -> vc=0.24
    # speed=85, limit=100 -> ratio=0.94 -> vc=0.20

    # Congestion-based car-following parameters
    if vc > 0.8:
        sigma = round(0.6 + (vc - 0.8) * 1.0, 2)
        tau = round(0.8 + (1.0 - vc) * 2.0, 1)
    elif vc > 0.5:
        sigma = round(0.4 + (vc - 0.5) * 0.67, 2)
        tau = round(1.0 + (0.8 - vc) * 1.67, 1)
    else:
        sigma = round(0.2 + vc * 0.4, 2)
        tau = round(1.5 + (0.5 - vc) * 2.0, 1)

    sigma = min(max(sigma, 0.1), 1.0)
    tau = min(max(tau, 0.5), 3.0)

    if "고속" in road_type:
        accel, decel, min_gap = 2.0, 3.5, 4.0
    else:
        accel, decel, min_gap = 2.6, 4.5, 2.5

    return {
        "sigma": sigma,
        "tau": tau,
        "accel": accel,
        "decel": decel,
        "minGap": min_gap,
        "speedFactor": round(speed_ratio, 2),
        "v_c_ratio": round(vc, 2),
    }


def estimate_vehicle_composition(road_type: str, hour: int) -> dict:
    """Estimate vehicle composition based on road type and time of day.

    Uses general traffic engineering standards since real data is unavailable.
    Can be replaced with real data later.
    """
    if "고속" in road_type:
        if 0 <= hour < 6:
            return {"passenger": 0.65, "truck": 0.32, "bus": 0.03}
        else:
            return {"passenger": 0.78, "truck": 0.17, "bus": 0.05}
    elif "간선" in road_type:
        return {"passenger": 0.83, "truck": 0.09, "bus": 0.08}
    else:
        return {"passenger": 0.90, "truck": 0.05, "bus": 0.05}


def get_speed_limit(road_type: str, area_type: str) -> int:
    """Estimate speed limit from road type and urban/suburban classification."""
    if "고속도로" in road_type:
        return 100 if "외곽" in area_type else 80
    elif "도시고속" in road_type or "자동차전용" in road_type:
        return 80
    elif "주간선" in road_type:
        return 60 if "외곽" in area_type else 50
    elif "보조간선" in road_type:
        return 50 if "외곽" in area_type else 50
    else:
        return 50


def load_nodelink_speeds(shp_dir: str = None) -> dict:
    """Load actual speed limits per link from the standard node-link SHP."""
    if shp_dir is None:
        raw_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "raw_data")
        import unicodedata
        for d in os.listdir(raw_dir):
            dn = unicodedata.normalize("NFC", d)
            if "NODELINKDATA" in dn:
                shp_dir = os.path.join(raw_dir, d)
                break

    if not shp_dir:
        return {}

    shp_path = os.path.join(shp_dir, "MOCT_LINK")
    if not os.path.exists(shp_path + ".shp"):
        return {}

    try:
        import shapefile
    except ImportError:
        print("  pyshp not found -> skipping node-link")
        return {}

    print(f"Loading node-link: {shp_path}")
    sf = shapefile.Reader(shp_path, encoding='euc-kr')

    # Load all node-link data (for matching via mapping)
    nodelink = {}
    for rec in sf.iterRecords():
        lid = str(rec['LINK_ID'])
        nodelink[lid] = {
            "max_spd": int(rec['MAX_SPD']) if rec['MAX_SPD'] else None,
            "lanes": int(rec['LANES']) if rec['LANES'] else None,
            "road_rank": str(rec['ROAD_RANK']),
            "length": float(rec['LENGTH']) if rec['LENGTH'] else 0,
        }

    print(f"  Node-link records: {len(nodelink)}")

    # Load TOPIS service link -> standard link mapping table
    mapping_file = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "raw_data", "TOPISto", "서울시 표준링크 매핑정보_2025년3월 기준.xlsx"
    )
    svc_to_std = {}
    if os.path.exists(mapping_file):
        wb_map = openpyxl.load_workbook(mapping_file, read_only=True)
        ws_map = wb_map[wb_map.sheetnames[0]]
        for i, row in enumerate(ws_map.iter_rows(values_only=True)):
            if i == 0:
                continue
            svc = str(int(row[0]))
            std = str(int(row[1]))
            if svc not in svc_to_std:
                svc_to_std[svc] = []
            svc_to_std[svc].append(std)
        wb_map.close()
        print(f"  Mapping table: {len(svc_to_std)} service links")

    return nodelink, svc_to_std


def resolve_link(svc_id, nodelink, svc_to_std):
    """Find node-link info by service link ID (directly or via mapping)."""
    if svc_id in nodelink:
        return nodelink[svc_id]
    if svc_id in svc_to_std:
        for std_id in svc_to_std[svc_id]:
            if std_id in nodelink:
                return nodelink[std_id]
    return None


def build_training_pairs(speed_data: dict, volume_data: dict, nodelink: dict = None, svc_to_std: dict = None) -> list:
    """Combine speed and volume data to create training data pairs."""
    if nodelink is None:
        nodelink = {}
    if svc_to_std is None:
        svc_to_std = {}
    pairs = []

    # 1) Speed data based (where road name matching is possible)
    for road_name, data in speed_data.items():
        info = data["info"]
        lanes = info.get("lanes", 2)
        road_type = info.get("road_type", "기타")
        area_type = info.get("area_type", "")

        # Extract measured info from node-link (via mapping table)
        link_ids = info.get("link_ids", set())
        matched_links = []
        for lid in link_ids:
            resolved = resolve_link(lid, nodelink, svc_to_std)
            if resolved:
                matched_links.append(resolved)

        # Speed limit: prioritize node-link measured value
        real_speed_limit = None
        for ml in matched_links:
            if ml["max_spd"]:
                real_speed_limit = ml["max_spd"]
                break
        speed_limit = real_speed_limit if real_speed_limit else get_speed_limit(road_type, area_type)
        speed_limit_source = "measured" if real_speed_limit else "estimated"

        # Geometry statistics (node-link)
        if matched_links:
            link_lengths = [ml["length"] for ml in matched_links if ml["length"] > 0]
            avg_block_m = int(sum(link_lengths) / len(link_lengths)) if link_lengths else 0
            total_road_length_m = int(sum(link_lengths))
            intersection_count = len(link_lengths)
            intersection_density = "높음" if avg_block_m < 200 else ("보통" if avg_block_m < 400 else "낮음")
            has_median = road_type in ("주간선도로", "도시고속도로")
        else:
            avg_block_m = 0
            total_road_length_m = 0
            intersection_count = 0
            intersection_density = ""
            has_median = False

        hourly = data["hourly_speed"]
        if not hourly:
            continue

        # Group by time period
        for (h_start, h_end), time_label in TIME_LABELS.items():
            if h_start <= h_end:
                hours_in_range = [h for h in range(h_start, h_end)]
            else:
                hours_in_range = [h for h in range(h_start, 24)] + [h for h in range(0, h_end)]

            speeds_in_range = [hourly[h] for h in hours_in_range if h in hourly]
            if not speeds_in_range:
                continue

            avg_speed = round(sum(speeds_in_range) / len(speeds_in_range), 1)

            # Attempt volume matching (search by road name)
            volume_vph = None
            for vol_name, vol_data in volume_data.items():
                # Check if road name is contained in volume point name
                road_base = road_name.split("(")[0].strip()
                if road_base in vol_name:
                    vol_hours = [vol_data["hourly_volume"].get(h, 0) for h in hours_in_range
                                 if h in vol_data["hourly_volume"]]
                    if vol_hours:
                        volume_vph = int(sum(vol_hours) / len(vol_hours))
                    break

            # Skip if no volume data (only use roads with measured volume match)
            if volume_vph is None:
                continue

            cf = estimate_cf_params(avg_speed, speed_limit, road_type)
            comp = estimate_vehicle_composition(road_type, hours_in_range[0] if hours_in_range else 12)

            int_lanes = int(lanes) if lanes else 2
            area_short = area_type if area_type else "서울"

            # Determine simulation extent (by road type)
            radius_m = {
                "고속도로": 1000, "도시고속도로": 800,
                "주간선도로": 500, "보조간선도로": 400,
            }.get(road_type, 300)

            # Generate natural language prompts (expressions users might actually type)
            total_lanes = int_lanes * 2  # Both directions
            congestion = "막히는" if cf["v_c_ratio"] > 0.7 else ("보통인" if cf["v_c_ratio"] > 0.4 else "한산한")
            road_type_short = road_type.replace("도로", "")  # 보조간선도로 -> 보조간선

            prompts = [
                # Using actual road name
                f"{road_name} {time_label} 시뮬레이션 해줘",
                f"{road_name} {time_label}",
                # Natural language description (without road name)
                f"{congestion} {area_short} {road_type_short} 왕복 {total_lanes}차선 {time_label}",
                f"왕복 {total_lanes}차선 {road_type_short} {time_label} 교통 시뮬레이션",
                # Mixed type
                f"{road_name} 같은 {road_type_short} {time_label} 상황",
            ]

            # Construct reasoning
            geo_desc = ""
            if avg_block_m > 0:
                geo_desc = (
                    f" 도로 총길이 {total_road_length_m}m, "
                    f"평균블록간격 {avg_block_m}m, "
                    f"교차로밀도 {intersection_density}."
                )

            params = {
                "speed_kmh": avg_speed,
                "volume_vph": volume_vph,
                "lanes": int_lanes,
                "speed_limit_kmh": speed_limit,
                "sigma": cf["sigma"],
                "tau": cf["tau"],
                "avg_block_m": avg_block_m if avg_block_m > 0 else int(radius_m * 0.4),
                "reasoning": (
                    f"{area_short} {road_type} 편도{int_lanes}차로 제한{speed_limit}km/h. "
                    f"{time_label} 평균{avg_speed}km/h V/C{cf['v_c_ratio']:.2f}. "
                    f"블록간격{avg_block_m}m."
                ),
            }

            for prompt in prompts:
                pairs.append({
                    "prompt": prompt,
                    "params": params,
                    "meta": {
                        "road": road_name,
                        "time": time_label,
                        "road_type": road_type,
                        "area_type": area_type,
                        "volume_source": "실측",
                        "source": "서울시 2025.10 실측",
                    },
                })

    return pairs


def to_openai_jsonl(pairs: list) -> list:
    """Convert to OpenAI fine-tuning JSONL format."""
    return [
        {
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": p["prompt"]},
                {"role": "assistant", "content": json.dumps(p["params"], ensure_ascii=False)},
            ]
        }
        for p in pairs
    ]


def to_gemini_jsonl(pairs: list) -> list:
    """Convert to Gemini tuning JSONL format."""
    return [
        {
            "text_input": f"[시스템] {SYSTEM_PROMPT}\n[사용자] {p['prompt']}",
            "output": json.dumps(p["params"], ensure_ascii=False),
        }
        for p in pairs
    ]


def save_jsonl(records: list, filepath: str):
    """Save records to a JSONL file."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"  Saved: {filepath} ({len(records)} records)")


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Find files
    speed_file = None
    volume_file = None
    import unicodedata
    for f in os.listdir(RAW_DIR):
        fpath = os.path.join(RAW_DIR, f)
        fn = unicodedata.normalize("NFC", f)
        if fn.startswith("temp"):
            continue
        if "속도" in fn or "통행속도" in fn:
            speed_file = fpath
        elif "교통량" in fn:
            volume_file = fpath

    if not speed_file:
        print("Speed data file not found.")
        print(f"  Please place the 'Seoul vehicle travel speed' file in raw_data/.")
        return

    print("=" * 50)
    print("  Real data-based fine-tuning dataset generation")
    print("=" * 50)

    # Load data
    speed_data = load_speed_data(speed_file)

    volume_data = {}
    if volume_file:
        volume_data = load_volume_data(volume_file)
    else:
        print("  No volume data -> using speed-based reverse estimation")

    # Node-link (measured speed limits + geometry)
    nodelink, svc_to_std = load_nodelink_speeds()

    # Generate training pairs
    pairs = build_training_pairs(speed_data, volume_data, nodelink, svc_to_std)
    print(f"\nTotal training pairs: {len(pairs)} records")

    # Shuffle
    random.seed(42)
    random.shuffle(pairs)

    # Train/val split (90/10)
    split = int(len(pairs) * 0.9)
    train_pairs = pairs[:split]
    val_pairs = pairs[split:]

    # Save
    print("\nSaving...")
    save_jsonl(to_openai_jsonl(train_pairs), os.path.join(OUTPUT_DIR, "train_real_openai.jsonl"))
    save_jsonl(to_openai_jsonl(val_pairs), os.path.join(OUTPUT_DIR, "val_real_openai.jsonl"))
    save_jsonl(to_gemini_jsonl(train_pairs), os.path.join(OUTPUT_DIR, "train_real_gemini.jsonl"))
    save_jsonl(pairs, os.path.join(OUTPUT_DIR, "train_real_raw.jsonl"))

    # Statistics
    speeds = [p["params"]["speed_kmh"] for p in pairs]
    volumes = [p["params"]["volume_vph"] for p in pairs]
    roads = set(p["meta"]["road"] for p in pairs)
    print(f"\nStatistics:")
    print(f"  Roads: {len(roads)}")
    print(f"  Speed range: {min(speeds):.1f} ~ {max(speeds):.1f} km/h")
    print(f"  Volume range: {min(volumes)} ~ {max(volumes)} vph")
    print(f"  Train: {len(train_pairs)} records, Validation: {len(val_pairs)} records")

    # Sample output
    print(f"\nSample 3 records:")
    for p in pairs[:3]:
        print(f"  Input: {p['prompt']}")
        print(f"  Speed: {p['params']['speed_kmh']}km/h, Volume: {p['params']['volume_vph']}vph")
        print()


if __name__ == "__main__":
    main()
